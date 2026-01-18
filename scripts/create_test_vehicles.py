#!/usr/bin/env python3
"""
Create Test Vehicle Database
Generates a sample vehicles.xlsx file for testing vehicle fuzzy matching.
"""

import openpyxl
from pathlib import Path


def create_test_excel():
    """Create test Excel file with sample vehicle data."""
    print("=" * 60)
    print("CREATING TEST VEHICLE DATABASE")
    print("=" * 60)

    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehicles"

    # Set headers
    ws['A1'] = 'id'
    ws['B1'] = 'legalName'

    # Make headers bold
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Sample test data with various company names
    # These are designed to test fuzzy matching at different similarity levels
    test_data = [
        ("V001", "Acme Corporation"),
        ("V002", "Global Logistics GmbH"),
        ("V003", "Tech Solutions Inc."),
        ("V004", "Deutsche Transportgesellschaft mbH"),
        ("V005", "Swift Shipping Ltd"),
        ("V006", "EuroFreight Systems AG"),
        ("V007", "Premier Auto Services"),
        ("V008", "International Trade Partners"),
        ("V009", "Blue Sky Industries"),
        ("V010", "Northern Logistics Solutions"),
        ("V011", "Advanced Manufacturing Co."),
        ("V012", "Pacific Trading Company"),
        ("V013", "Metro Delivery Services"),
        ("V014", "Continental Express Transport"),
        ("V015", "United Distribution Network"),
    ]

    # Populate data
    print(f"\nAdding {len(test_data)} vehicles to database:")
    print("-" * 60)

    for i, (vehicle_id, legal_name) in enumerate(test_data, start=2):
        ws[f'A{i}'] = vehicle_id
        ws[f'B{i}'] = legal_name
        print(f"  {vehicle_id:6} | {legal_name}")

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40

    # Create output directory if it doesn't exist
    output_path = Path("vehicles/vehicles.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Save workbook
    wb.save(output_path)

    print("-" * 60)
    print(f"\n✓ Successfully created: {output_path}")
    print(f"  Sheet name: {ws.title}")
    print(f"  Columns: id, legalName")
    print(f"  Rows: {len(test_data)} vehicles")
    print("\n" + "=" * 60)
    print("TEST DATABASE READY")
    print("=" * 60)
    print("\nTest matching examples:")
    print("  - 'Acme Corp' → should match 'Acme Corporation' (V001)")
    print("  - 'Global Logistics' → should match 'Global Logistics GmbH' (V002)")
    print("  - 'Tech Solutions' → should match 'Tech Solutions Inc.' (V003)")
    print("  - 'Random Company' → should return N/A (no match)")
    print("\nRun the application to test vehicle matching!")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    try:
        create_test_excel()
    except Exception as e:
        print(f"\n✗ Error creating test database: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
