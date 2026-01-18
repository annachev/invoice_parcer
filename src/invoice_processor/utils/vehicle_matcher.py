#!/usr/bin/env python3
"""
Vehicle Matcher
Fuzzy matching of invoice recipient names against a vehicle database.
"""

from pathlib import Path
from typing import Dict, Tuple, Optional
import openpyxl
from rapidfuzz import fuzz

from ..core.logging_config import get_logger
from ..core.exceptions import ConfigurationError

logger = get_logger(__name__)


class VehicleMatcher:
    """
    Fuzzy matcher for matching invoice recipients to vehicle database.

    Features:
    - Lazy loading (load Excel on first use)
    - Configurable similarity threshold (default 90%)
    - Caching of vehicle database in memory
    - Graceful error handling (missing file disables matching)

    Usage:
        matcher = VehicleMatcher("vehicles/vehicles.xlsx", threshold=0.9)
        matched_name, vehicle_id, score = matcher.match_recipient("Acme Corp")

        if score >= 0.9:
            # Use matched_name and vehicle_id
            print(f"Matched: {matched_name} (ID: {vehicle_id})")
        else:
            # Keep original, show N/A
            print(f"No match found (score: {score:.2f})")
    """

    def __init__(self, excel_path: str, threshold: float = 0.9):
        """
        Initialize vehicle matcher.

        Args:
            excel_path: Path to Excel file with Vehicles sheet
            threshold: Similarity threshold (0.0-1.0, default 0.9)
        """
        self.excel_path = Path(excel_path)
        self.threshold = threshold
        self.vehicles: Optional[Dict[str, str]] = None  # {id: legalName}
        self.enabled = True

        # Validate threshold
        if not 0.0 <= threshold <= 1.0:
            logger.warning(f"Invalid threshold {threshold}, using default 0.9")
            self.threshold = 0.9

    def load_vehicles(self) -> Dict[str, str]:
        """
        Load vehicle database from Excel file.

        Returns:
            Dictionary mapping vehicle_id -> legalName

        Raises:
            ConfigurationError: If Excel file is invalid or missing required structure
        """
        if not self.excel_path.exists():
            error_msg = f"Vehicle database not found: {self.excel_path}"
            logger.error(error_msg)
            self.enabled = False
            raise ConfigurationError(error_msg)

        try:
            logger.info(f"Loading vehicle database from: {self.excel_path}")

            # Load workbook
            workbook = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)

            # Validate sheet exists
            if "Vehicles" not in workbook.sheetnames:
                error_msg = f"Sheet 'Vehicles' not found in {self.excel_path}. Available sheets: {workbook.sheetnames}"
                logger.error(error_msg)
                self.enabled = False
                raise ConfigurationError(error_msg)

            sheet = workbook["Vehicles"]

            # Validate headers (first row)
            headers = [cell.value for cell in sheet[1]]

            if "id" not in headers or "legalName" not in headers:
                error_msg = f"Missing required columns in Vehicles sheet. Expected 'id' and 'legalName', found: {headers}"
                logger.error(error_msg)
                self.enabled = False
                raise ConfigurationError(error_msg)

            # Find column indices
            id_col = headers.index("id")
            name_col = headers.index("legalName")

            # Load data
            vehicles = {}
            row_count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                if not row or not row[id_col] or not row[name_col]:
                    continue  # Skip empty rows

                vehicle_id = str(row[id_col]).strip()
                legal_name = str(row[name_col]).strip()

                if vehicle_id and legal_name:
                    vehicles[vehicle_id] = legal_name
                    row_count += 1

            workbook.close()

            logger.info(f"Loaded {row_count} vehicles from database")

            if row_count == 0:
                logger.warning("Vehicle database is empty")

            return vehicles

        except openpyxl.utils.exceptions.InvalidFileException as e:
            error_msg = f"Invalid Excel file: {self.excel_path} - {e}"
            logger.error(error_msg)
            self.enabled = False
            raise ConfigurationError(error_msg) from e

        except Exception as e:
            error_msg = f"Failed to load vehicle database: {e}"
            logger.error(error_msg)
            self.enabled = False
            raise ConfigurationError(error_msg) from e

    def match_recipient(self, recipient_name: str) -> Tuple[str, str, float]:
        """
        Match recipient name against vehicle database using fuzzy matching.

        Args:
            recipient_name: Recipient name from invoice

        Returns:
            Tuple of (matched_legal_name, vehicle_id, similarity_score)
            - If score >= threshold: Returns matched name and ID
            - If score < threshold: Returns original name, "N/A", score

        Example:
            matched_name, vehicle_id, score = matcher.match_recipient("Acme Corp")
            # ("Acme Corporation", "V001", 0.95) if match found
            # ("Acme Corp", "N/A", 0.75) if no match
        """
        # Lazy load vehicles on first use
        if self.vehicles is None:
            try:
                self.vehicles = self.load_vehicles()
            except ConfigurationError:
                # Loading failed, matching disabled
                self.enabled = False
                return (recipient_name, "N/A", 0.0)

        # If disabled or no vehicles, return N/A
        if not self.enabled or not self.vehicles:
            return (recipient_name, "N/A", 0.0)

        # If recipient name is empty or invalid, return N/A
        if not recipient_name or recipient_name.strip() == "" or recipient_name == "PARSING FAILED":
            return (recipient_name, "N/A", 0.0)

        # Normalize recipient name for matching (case-insensitive)
        recipient_normalized = recipient_name.strip()

        # Find best match using fuzzy matching
        best_match_id = None
        best_match_name = None
        best_score = 0.0

        for vehicle_id, legal_name in self.vehicles.items():
            # Use rapidfuzz.fuzz.ratio for fuzzy string matching (0-100 scale)
            # This is similar to Levenshtein distance ratio
            score_percent = fuzz.ratio(recipient_normalized.lower(), legal_name.lower())
            score = score_percent / 100.0  # Convert to 0.0-1.0 scale

            if score > best_score:
                best_score = score
                best_match_id = vehicle_id
                best_match_name = legal_name

        # Check if best match meets threshold
        if best_score >= self.threshold and best_match_id:
            logger.debug(f"Vehicle match: '{recipient_name}' → '{best_match_name}' (ID: {best_match_id}, score: {best_score:.2f})")
            return (best_match_name, best_match_id, best_score)
        else:
            logger.debug(f"No vehicle match for '{recipient_name}' (best score: {best_score:.2f}, threshold: {self.threshold:.2f})")
            return (recipient_name, "N/A", best_score)

    def reload(self):
        """
        Reload vehicle database from Excel file.

        Useful for refreshing data if the Excel file has been updated.
        """
        self.vehicles = None
        self.enabled = True
        logger.info("Vehicle database cache cleared, will reload on next match")


if __name__ == "__main__":
    # Test vehicle matcher
    import sys

    print("=" * 60)
    print("VEHICLE MATCHER TEST")
    print("=" * 60)

    # Test with sample data
    excel_path = "vehicles/vehicles.xlsx"

    if not Path(excel_path).exists():
        print(f"\n✗ Excel file not found: {excel_path}")
        print("Run 'python scripts/create_test_vehicles.py' to create test data")
        sys.exit(1)

    print(f"\nLoading vehicles from: {excel_path}")

    try:
        matcher = VehicleMatcher(excel_path, threshold=0.9)

        # Test cases
        test_cases = [
            "Acme Corporation",      # Exact match (100%)
            "Acme Corp",             # High fuzzy match (~95%)
            "Global Logistics",      # Partial match (~85%)
            "Random Company XYZ",    # No match (<50%)
            "Tech Solutions",        # Fuzzy match (~90%)
        ]

        print(f"\nThreshold: {matcher.threshold:.0%}")
        print("\nTest Cases:")
        print("-" * 60)

        for test_name in test_cases:
            matched_name, vehicle_id, score = matcher.match_recipient(test_name)

            status = "✓ MATCHED" if score >= matcher.threshold else "✗ NO MATCH"
            print(f"\n{status}")
            print(f"  Input:    '{test_name}'")
            print(f"  Match:    '{matched_name}'")
            print(f"  Vehicle:  {vehicle_id}")
            print(f"  Score:    {score:.2%}")

        print("\n" + "=" * 60)
        print("TEST COMPLETED")
        print("=" * 60)

    except ConfigurationError as e:
        print(f"\n✗ Configuration error: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"\n✗ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
